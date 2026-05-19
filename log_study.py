#!/usr/bin/env python3
"""
Auto Study Logger — updated for new tracker layout
Usage: python3 log_study.py
Paste your daily log when prompted, press Enter twice when done.
Script appends to Excel, commits to GitHub automatically.

New Daily Log columns (updated tracker):
Col 1:  Day           ← auto-filled (pre-numbered in tracker)
Col 2:  Date
Col 3:  Month
Col 4:  Track
Col 5:  Topics Covered
Col 6:  Hours
Col 7:  SQL Problems
Col 8:  Python Article
Col 9:  Krish Naik
Col 10: DSA Problem
Col 11: GitHub ✓
Col 12: Stuck On
Col 13: Key Insight / Lesson
"""

import os
import re
from datetime import datetime
from pathlib import Path

# ── CONFIG — update these paths ──────────────────────────────
EXCEL_PATH = Path.home() / "Desktop" / "Andrina-study-journey" / "Study_Progress_Tracker.xlsx"
REPO_PATH  = Path.home() / "Desktop" / "Andrina-study-journey"
# ─────────────────────────────────────────────────────────────


def parse_log(text):
    t = text.lower()

    # Date
    date = datetime.today().strftime('%b %d %Y')

    # Month
    month = 'May — Month 1'
    month_map = {
        1: 'May — Month 1',
        2: 'Jun — Month 2',
        3: 'Jul — Month 3',
        4: 'Aug — Month 4',
        5: 'Sep — Month 5',
        6: 'Oct — Month 6',
        7: 'Nov — Month 7',
        8: 'Dec — Month 8',
    }
    for i in range(8, 0, -1):
        if f'month {i}' in t:
            month = month_map[i]
            break

    # Track
    track = 'DE'
    if 'de+ds' in t or ('de' in t and 'ds' in t):
        track = 'DE+DS'
    elif 'agentic' in t or ' ai' in t:
        track = 'AI'
    elif 'ds' in t or 'data science' in t or 'ml' in t:
        track = 'DS'
    elif 'de' in t or 'data engineering' in t:
        track = 'DE'

    # Topics Covered
    topics = ''
    for pattern in [r'topics?[:\s]+([^,\n]+)', r'studied[:\s]+([^,\n]+)', r'covered[:\s]+([^,\n]+)']:
        m = re.search(pattern, t)
        if m:
            topics = m.group(1).strip()
            break

    # Hours
    hours = 0
    m = re.search(r'(\d+\.?\d*)\s*h(?:our|r)?s?', t)
    if m:
        hours = float(m.group(1))

    # SQL Problems
    sql = 0
    m = re.search(r'(\d+)\s*sql', t)
    if m:
        sql = int(m.group(1))

    # Python Article — did they read/code a DataVidhya article?
    python_article = ''
    if any(w in t for w in ['python article yes', 'article yes', 'datavidhya yes', 'read article']):
        python_article = 'Yes'
    elif any(w in t for w in ['python article no', 'article no', 'no article']):
        python_article = 'No'

    # Krish Naik
    krish = ''
    if any(w in t for w in ['krish yes', 'krish naik yes', 'attended krish', 'watched krish']):
        krish = 'Yes'
    elif any(w in t for w in ['krish no', 'no krish']):
        krish = 'No'

    # DSA Problem — which problem did they do?
    dsa = ''
    m = re.search(r'dsa[:\s]+([^,\.\n]+)', t)
    if m:
        dsa = m.group(1).strip()
    else:
        m = re.search(r'lc\s*\d+[^,\.\n]*', t)
        if m:
            dsa = m.group(0).strip()
        elif 'no dsa' in t or 'dsa no' in t:
            dsa = 'None'

    # GitHub
    github = ''
    if any(w in t for w in ['github yes', 'committed', 'pushed', 'git yes']):
        github = 'Yes'
    elif any(w in t for w in ['github no', 'no github']):
        github = 'No'

    # Stuck On
    stuck = ''
    m = re.search(r'stuck\s+on[:\s]+([^,\.\n]+)', t)
    if m:
        stuck = m.group(1).strip()

    # Key Insight / Lesson
    insight = ''
    for pattern in [r'insight[:\s]+([^,\.\n]+)', r'learned[:\s]+([^,\.\n]+)', r'clicked[:\s]+([^,\.\n]+)']:
        m = re.search(pattern, t)
        if m:
            insight = m.group(1).strip()
            break

    return {
        'date':           date,
        'month':          month,
        'track':          track,
        'topics':         topics.title() if topics else '',
        'hours':          hours,
        'sql':            sql,
        'python_article': python_article,
        'krish':          krish,
        'dsa':            dsa.title() if dsa else '',
        'github':         github,
        'stuck':          stuck.title() if stuck else '',
        'insight':        insight.title() if insight else '',
    }


def append_to_excel(data):
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Installing openpyxl...")
        os.system("pip install openpyxl --break-system-packages -q")
        from openpyxl import load_workbook

    if not EXCEL_PATH.exists():
        print(f"\n  Excel file not found at:\n  {EXCEL_PATH}")
        print("  Please check the EXCEL_PATH in the script config.")
        return False

    wb = load_workbook(EXCEL_PATH)
    ws = wb['Daily Log']

    # Find first row where Date (col 2) is empty but Day (col 1) is pre-filled
    target_row = None
    for row in ws.iter_rows(min_row=3, max_col=2):
        day_val  = row[0].value
        date_val = row[1].value
        if day_val is not None and date_val is None:
            target_row = row[0].row
            break

    if target_row is None:
        print("  No empty rows found in Daily Log — tracker may be full.")
        return False

    # Write to correct columns (1-indexed)
    # Col 1 = Day (already pre-filled — skip)
    ws.cell(row=target_row, column=2,  value=data['date'])           # Date
    ws.cell(row=target_row, column=3,  value=data['month'])          # Month
    ws.cell(row=target_row, column=4,  value=data['track'])          # Track
    ws.cell(row=target_row, column=5,  value=data['topics'])         # Topics Covered
    ws.cell(row=target_row, column=6,  value=data['hours'])          # Hours
    ws.cell(row=target_row, column=7,  value=data['sql'])            # SQL Problems
    ws.cell(row=target_row, column=8,  value=data['python_article']) # Python Article
    ws.cell(row=target_row, column=9,  value=data['krish'])          # Krish Naik
    ws.cell(row=target_row, column=10, value=data['dsa'])            # DSA Problem
    ws.cell(row=target_row, column=11, value=data['github'])         # GitHub ✓
    ws.cell(row=target_row, column=12, value=data['stuck'])          # Stuck On
    ws.cell(row=target_row, column=13, value=data['insight'])        # Key Insight / Lesson

    wb.save(EXCEL_PATH)
    print(f"  Excel updated — row {target_row} written")
    return True


def git_commit(data):
    os.chdir(REPO_PATH)
    msg = f"Study log {data['date']} — {data['track']} — {data['hours']}hrs — {data['sql']} SQL"
    os.system(f'git add "{EXCEL_PATH.name}"')
    os.system(f'git commit -m "{msg}"')
    os.system('git push')
    print(f"  GitHub committed: {msg}")


def evaluate_progress(data):
    print("\n" + "=" * 55)
    print("  PROGRESS EVALUATION")
    print("=" * 55)

    warnings  = []
    positives = []

    # Hours
    if data['hours'] >= 6:
        positives.append(f"{data['hours']} hrs studied — on target")
    elif data['hours'] >= 4:
        warnings.append(f"{data['hours']} hrs — target is 6-8 hrs per day")
    elif data['hours'] >= 2:
        warnings.append(f"Only {data['hours']} hrs — below 6hr target")
    else:
        warnings.append(f"Only {data['hours']} hrs — needs urgent attention tomorrow")

    # SQL
    if data['sql'] >= 2:
        positives.append(f"{data['sql']} SQL problems — daily habit maintained")
    elif data['sql'] == 1:
        warnings.append("Only 1 SQL problem — target is 2 minimum every day")
    else:
        warnings.append("No SQL problems today — do 4 tomorrow to catch up")

    # Python Article
    if data['python_article'] == 'Yes':
        positives.append("DataVidhya Python article read + recoded")
    elif data['python_article'] == 'No':
        warnings.append("No Python article today — read + recode tomorrow")

    # Krish Naik
    if data['krish'] == 'Yes':
        positives.append("Krish Naik session watched + recoded")
    elif data['krish'] == 'No':
        warnings.append("No Krish Naik today — stay on catch-up schedule")

    # DSA
    if data['dsa'] and data['dsa'].lower() != 'none':
        positives.append(f"DSA done — {data['dsa']}")
    else:
        warnings.append("No DSA today — daily habit broken, do it first tomorrow")

    # GitHub
    if data['github'] == 'Yes':
        positives.append("GitHub committed — green square earned")
    else:
        warnings.append("No GitHub commit — push something before midnight")

    for p in positives:
        print(f"  + {p}")
    for w in warnings:
        print(f"  ! {w}")

    print()
    if not warnings:
        print("  Perfect day Anna. This is what 8 months of")
        print("  consistency looks like. Keep this up.")
    elif len(warnings) == 1:
        print(f"  Strong day overall. One thing to fix tomorrow:")
        print(f"  {warnings[0]}")
    elif len(warnings) == 2:
        print("  Decent day. Two things to tighten tomorrow.")
    else:
        print("  Below target today. Tomorrow:")
        print("  SQL first. Then DSA. No exceptions.")

    if data['stuck']:
        print()
        print(f"  Stuck on: {data['stuck']}")
        print("  Ask Claude: 'explain [topic] with a biochemistry")
        print("  analogy and a code example under 20 lines'")

    print()
    print(f"  At 6 hrs/day — you finish 8 months in Dec 2026.")
    print(f"  Today: {data['hours']} hrs.")
    print("=" * 55)


def main():
    print("\n" + "=" * 55)
    print("  ANNA'S DAILY STUDY LOG")
    print("  Target: 6-8 hrs/day · 2 SQL · DSA · 1 GitHub commit")
    print("=" * 55)
    print("\nPaste your log below (press Enter twice when done):\n")
    print("Example:")
    print("  May 18, Month 1, DE, topics: pandas groupby merge,")
    print("  6 hours, 2 SQL, article yes, krish yes,")
    print("  dsa: LC 1 Two Sum, github yes,")
    print("  stuck on: merge vs join,")
    print("  insight: groupby in pandas = GROUP BY in SQL")
    print()

    lines = []
    while True:
        try:
            line = input()
            if line == '' and lines and lines[-1] == '':
                break
            lines.append(line)
        except EOFError:
            break

    raw_text = ' '.join(lines)
    if not raw_text.strip():
        print("No log entered. Exiting.")
        return

    print("\nParsing your log...")
    data = parse_log(raw_text)

    print("\nParsed values:")
    print(f"  Date:           {data['date']}")
    print(f"  Month:          {data['month']}")
    print(f"  Track:          {data['track']}")
    print(f"  Topics:         {data['topics']}")
    print(f"  Hours:          {data['hours']}")
    print(f"  SQL Problems:   {data['sql']}")
    print(f"  Python Article: {data['python_article']}")
    print(f"  Krish Naik:     {data['krish']}")
    print(f"  DSA Problem:    {data['dsa']}")
    print(f"  GitHub:         {data['github']}")
    print(f"  Stuck On:       {data['stuck']}")
    print(f"  Insight:        {data['insight']}")

    confirm = input("\nLooks correct? (y/n): ").strip().lower()
    if confirm != 'y':
        print("Log cancelled. Run again and re-enter.")
        return

    print("\nWriting to Excel...")
    success = append_to_excel(data)

    if success:
        print("Committing to GitHub...")
        git_commit(data)
        evaluate_progress(data)
        print("\nDone. See you tomorrow Anna.\n")
    else:
        print("Excel write failed — check EXCEL_PATH in script config.")


if __name__ == '__main__':
    main()
